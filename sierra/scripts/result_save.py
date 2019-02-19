#!/usr/bin/python3
import sys
from openpyxl import load_workbook, Workbook
import os
from openpyxl.styles import Font, colors, Alignment
import time

# get yocto version
YOCTO_VERSION=sys.argv[1]
MODEM_VERSION=sys.argv[2]
DIFF_TIME=sys.argv[3]
MNT_DIR=sys.argv[4]
LOG_FILE=sys.argv[5]

mnt=MNT_DIR
def get_filename(yocto_version, modem_version):
    modem_version=modem_version.strip()
    AR758x_pos = yocto_version.find("AR758x")
    AR759x_pos = yocto_version.find("AR759x")
    yocto_2_2_pos = yocto_version.find("LXSWI2.2")
    yocto_1_7_pos = yocto_version.find("LXSWI1.7")
    file_path = mnt + "/Misc/Public-Files/Public-Files-1/anson/LTP-test-report"
    if AR759x_pos >= 0 and yocto_2_2_pos >= 0:
        file_path = mnt + "/Engineering/Project_Management/Emu_AR759x_R2/01_Engineering/01_Software/05_Test_Report"
        file_name = file_path + "/SWI9X40_Test_report_" + modem_version + "-" + yocto_version + ".xlsx"
    elif AR758x_pos >= 0 and yocto_2_2_pos >= 0:
        file_path = mnt + "/Engineering/Project_Management/Eel_AR758x_R2/01_Engineering/01_Software/11_Test_Report"
        file_name = file_path + "/SWI9X28_Test_report_" + modem_version + "-" + yocto_version + ".xlsx"
    elif AR759x_pos >= 0 and yocto_1_7_pos >= 0:
        file_path = mnt + "/Engineering/Project_Management/Coronado_AR759x/01_Engineering/01_FW/10_Regression"
        file_name = file_path + "/SWI9X40_Test_report_" + modem_version + "-" + yocto_version + ".xlsx"
    elif AR758x_pos >= 0 and yocto_1_7_pos >= 0:
        file_path = mnt + "/Engineering/Project_Management/Cougar_AR758x/01_Engineering/01_Software/11_Test_Report"
        file_name = file_path + "/SWI9X28_Test_report_" + modem_version + "-" + yocto_version + ".xlsx"
    else:
        file_name = "invalid"
    return file_name


# get excel file name
file_name=get_filename(YOCTO_VERSION, MODEM_VERSION)
if file_name == 'invalid':
    print("invalid yocto version: " + YOCTO_VERSION)
    sys.exit(1)
print(file_name)

# mount to //cnshz-nv-f101/file
if os.path.exists(mnt) == False:
    os.mkdir(mnt)

os.system("echo 123456 | sudo -S umount -fl " + mnt)
os.system('echo 123456 | sudo -S mount -t cifs -o username="aouyang",password="lfwOYXF520221" //10.22.22.1/file ' + mnt)

if os.path.exists(file_name) == True:
    wb = load_workbook(file_name)
else:
    wb = Workbook()
sheet = wb.create_sheet('ltp')

# set colnum width
sheet.column_dimensions['A'].width=35
sheet.column_dimensions['B'].width=10
sheet.column_dimensions['C'].width=10

# color
font=Font('宋体', bold=True, color=colors.RED)

cur_date=time.localtime()
log_file=LOG_FILE
if os.path.exists(log_file):
    f=open(log_file)
    lines = f.readlines()
    length = len(lines)
    row=["Test reports("+str(cur_date.tm_year)+"-"+str(cur_date.tm_mon)+"-"+str(cur_date.tm_mday)+")"]
    sheet.append(row)
    row=["Elapsed time: "+DIFF_TIME + " minutes"]
    sheet.append(row)
    row=["Tag: "+YOCTO_VERSION]
    sheet.append(row)
    # test cases statistics summary
    for index in range(length-7, length-4):
        line = [lines[index]]
        sheet.append(line)
    row=[" "]
    sheet.append(row)
    for index in range(length-4, length-1):
        line = [lines[index]]
        sheet.append(line)
    # detail
    row=[" "]
    sheet.append(row)
    row=["---------------------------------", "----------", "----------"]
    sheet.append(row)
    row=["Detail"]
    sheet.append(row)
    header=['TestCases', 'ExitValue', 'Result']
    sheet.append(header)
    # each case record
    for index in range(4, length-8):
        line = lines[index]
        row=line.split()
        sheet.append(row)
    f.close()

# set color
sheet.cell(row=1, column=1).font=font
sheet.cell(row=13, column=1).font=font

# save excel
wb.save(file_name)
os.system("echo 123456 | sudo -S umount -fl " + mnt)

